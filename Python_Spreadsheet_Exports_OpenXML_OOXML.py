# Python Spreadsheet Exports.
# Exporting your Data as CSV, Excel (XLS), or XLSX in python with django.
#

#
# Office Open XML Format.
# XLSX (a.k.a. OOXML or OpenXML) is a zipped, XML-based file format developed by Microsoft.
# It is fully supported by Microsoft Office 2007 and newer versions.
# OpenOffice 4.0, for example, can only read it.
# There is a python library openpyxl for reading and writing those files.
# This format is great when you need more than 256 columns and text formatting options.
#

#
# Here, todo_obj is data source. 
#


def export_xlsx(request):

    import openpyxl
    from openpyxl.cell import get_column_letter

    todo_obj=Todo.objects.filter(user_id=request.session['user_id'])

    response = HttpResponse(mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=todo.xlsx'

    wb = openpyxl.Workbook()

    ws = wb.get_active_sheet()
    ws.title = "Todo"

    row_num = 0

    columns = [
        (u"sl", 15),
        (u"job", 100),
        (u"Date", 70),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)

        c.value = columns[col_num][0]

        # set column width

        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for obj in todo_obj:
        row_num += 1

        row = [
            row_num,
            obj.todo_job,
            obj.created_date.strftime("%A %d. %B %Y"),
        ]

        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)

            c.value = row[col_num]

    wb.save(response)

    return response 
